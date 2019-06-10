#coding=utf-8

def task_00():
    '''
        由于 office 自动调整行高的功能，导致文字上下离线太近，
        而每行的高度都不同，
        需求： 把各行高度都增加 25 像素
    '''
    from openpyxl import load_workbook
    from os import system as command
    
    # 载入 workbook, 显示 sheet 的名称
    wb = load_workbook('target.xlsx'); print(wb.sheetnames)

    # 选择 第2张 sheet
    ws = wb[wb.sheetnames[1]]

    # 6~64行， 每行高度增加 25；
    for i in range(6, 65):
        ws.row_dimensions[i].height += 25
    
    # 保存
    wb.save('targetX.xlsx')

    # 打开目录
    command("explorer .")

def task_01():
    '''
        由于 office 自动调整行高的功能，导致文字上下离线太近，
        而每行的高度都不同，
        需求： 把各行高度都增加 25 像素
    '''
    from openpyxl import load_workbook
    from os import system as command
    
    # 载入 workbook, 显示 sheet 的名称
    wb = load_workbook('target.xlsx')

    # 选择 第2张 sheet
    ws = wb[wb.sheetnames[1]]

    # 第5行所有数值
    for cell in ws['5']:
        print(cell.value)

    # 特定单元格数值
    print(ws['D6'].value)

def task_02():
    from openpyxl import Workbook
    # 创建一个工作表，取得活动页, 改名
    wb = Workbook()
    sheet = wb.active
    sheet.title = "New Sheet"
    # 这个属性是可读可写的。当然，这个只针对当前活动页，别的页的话，可以用create_sheet和remove_sheet进行添加和删除。
    # 往sheet页里面写内容就比较简单了，跟上面读一样，

    sheet['C3'] = 'Hello world!'
    for i in range(10):
        sheet["A%d" % (i+1)].value = i + 1

    # 我们还可以进行花式操作，比如写写公式：
    sheet["E1"].value = "=SUM(A:A)"

    # 最后记得保存
    wb.save('保存一个新的excel.xlsx')

def tmp():
    from openpyxl import Workbook
    import datetime

    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    ws['A4'] = datetime.datetime.now()
    ws.append([1, 2, 3, 4, 5, 8, 13])

    # Save the file
    wb.save("sample.xlsx")

if __name__ == "__main__":
    tmp()